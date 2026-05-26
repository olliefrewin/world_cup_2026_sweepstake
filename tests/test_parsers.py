# Tests for Part 1 and Part 2 parsers, and the classifier.

from __future__ import annotations

import pytest

from sweepstake.parsers.classifier import classify, ClassifierError
from sweepstake.parsers.part1 import parse as parse_part1
from sweepstake.parsers.part2 import parse as parse_part2
from sweepstake.constants import GROUPS


# ---------------------------------------------------------------------------
# Classifier
# ---------------------------------------------------------------------------

def test_classify_part1(valid_part1_path):
    assert classify(valid_part1_path) == "part1"


def test_classify_part2(valid_part2_path):
    assert classify(valid_part2_path) == "part2"


def test_classify_unknown_raises(unknown_sheet_path):
    with pytest.raises(ClassifierError):
        classify(unknown_sheet_path)


# ---------------------------------------------------------------------------
# Part 1 parser -- happy path
# ---------------------------------------------------------------------------

def test_part1_valid_parses_name(valid_part1_path):
    result = parse_part1(valid_part1_path)
    assert result.participant_name == "Alice"


def test_part1_valid_parses_all_groups(valid_part1_path):
    result = parse_part1(valid_part1_path)
    assert len(result.group_picks) == 12
    for g, (winner, runner) in result.group_picks.items():
        assert winner == GROUPS[g][0]
        assert runner == GROUPS[g][1]


def test_part1_valid_parses_finalists(valid_part1_path):
    result = parse_part1(valid_part1_path)
    assert result.finalist_1 == "Brazil"
    assert result.finalist_2 == "England"


def test_part1_valid_parses_winner(valid_part1_path):
    result = parse_part1(valid_part1_path)
    assert result.winner == "Brazil"


def test_part1_valid_parses_golden_boot(valid_part1_path):
    result = parse_part1(valid_part1_path)
    assert result.golden_boot_raw == "Lionel Messi"


def test_part1_valid_is_valid(valid_part1_path):
    result = parse_part1(valid_part1_path)
    assert result.is_valid


# ---------------------------------------------------------------------------
# Part 1 parser -- validation errors
# ---------------------------------------------------------------------------

def test_part1_invalid_group_pick_has_error(part1_invalid_group_pick_path):
    result = parse_part1(part1_invalid_group_pick_path)
    assert not result.is_valid
    assert any("Group A" in e for e in result.errors)


def test_part1_blank_name_has_error(part1_blank_name_path):
    result = parse_part1(part1_blank_name_path)
    assert not result.is_valid
    assert any("name" in e.lower() for e in result.errors)


def test_part1_duplicate_finalist_is_still_valid(part1_duplicate_finalist_path):
    # Duplicate finalists are not a validation error -- they just score at most once.
    # The scoring engine handles the deduplication.
    result = parse_part1(part1_duplicate_finalist_path)
    assert result.is_valid
    assert result.finalist_1 == "Brazil"
    assert result.finalist_2 == "Brazil"


# ---------------------------------------------------------------------------
# Part 2 parser -- happy path
# ---------------------------------------------------------------------------

def test_part2_valid_parses_name(valid_part2_path):
    result = parse_part2(valid_part2_path)
    assert result.participant_name == "Alice"


def test_part2_valid_has_all_r32_slots(valid_part2_path):
    result = parse_part2(valid_part2_path)
    assert len(result.r32_winners) == 16


def test_part2_valid_parses_tiebreaker(valid_part2_path):
    result = parse_part2(valid_part2_path)
    assert result.tiebreaker_final_goals == 3


def test_part2_valid_is_valid(valid_part2_path):
    result = parse_part2(valid_part2_path)
    assert result.is_valid


# ---------------------------------------------------------------------------
# Part 2 parser -- validation
# ---------------------------------------------------------------------------

def test_part2_all_blank_has_error(tmp_path):
    import openpyxl
    from sweepstake.constants import SHEET_PART2
    wb = openpyxl.Workbook()
    wb.active.title = "_tmp"
    ws = wb.create_sheet(SHEET_PART2)
    ws["B4"] = "Bob"
    # Leave all team cells blank
    p = tmp_path / "blank_part2.xlsx"
    wb.save(str(p))
    result = parse_part2(p)
    assert not result.is_valid
    assert any("blank" in e.lower() for e in result.errors)
