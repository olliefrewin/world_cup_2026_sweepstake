# Shared fixtures for the test suite.
# Synthetic .xlsx files are generated here so the suite has no hard dependency on
# binary fixture files for the common cases.

from __future__ import annotations

import io
import pytest
import openpyxl

from sweepstake.constants import GROUPS, SHEET_PART1, SHEET_PART2


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_part1_wb(
    name: str = "Alice",
    group_picks: dict | None = None,
    finalist_1: str | None = "Brazil",
    finalist_2: str | None = "England",
    winner: str | None = "Brazil",
    golden_boot: str = "Lionel Messi",
    submitted_on: str = "2026-06-01",
) -> openpyxl.Workbook:
    if group_picks is None:
        # Default: pick the first team as winner, second as runner-up for every group
        group_picks = {g: (GROUPS[g][0], GROUPS[g][1]) for g in GROUPS}

    wb = openpyxl.Workbook()
    wb.active.title = "_temp"
    ws = wb.create_sheet(SHEET_PART1)

    ws["B4"] = name
    ws["B47"] = submitted_on

    # Group picks -- rows 23-34
    group_rows = {g: 23 + i for i, g in enumerate(GROUPS)}
    for g, (win, run) in group_picks.items():
        row = group_rows[g]
        ws[f"C{row}"] = win
        ws[f"D{row}"] = run

    ws["B40"] = finalist_1
    ws["B41"] = finalist_2
    ws["B42"] = winner
    ws["B43"] = golden_boot

    return wb


def _make_part2_wb(
    name: str = "Alice",
    r32_teams: dict | None = None,
    r32_winners: dict | None = None,
    r16_winners: dict | None = None,
    qf_winners: dict | None = None,
    sf_winners: dict | None = None,
    champion: str | None = "Brazil",
    tiebreaker: int | None = 3,
) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.active.title = "_temp"
    ws = wb.create_sheet(SHEET_PART2)

    ws["B4"] = name

    # Default R32 left teams (col A) and winners (col B)
    left_rows = [8, 11, 14, 17, 20, 23, 26, 29]
    teams_cycle = list(GROUPS["A"]) * 10  # enough teams for defaults

    for i, row in enumerate(left_rows):
        team_a = teams_cycle[i * 2]
        team_b = teams_cycle[i * 2 + 1]
        ws[f"A{row}"] = team_a
        ws[f"A{row + 1}"] = team_b
        ws[f"B{row}"] = team_a  # winner = first team by default

    # Right side (col K teams, col J winners)
    right_rows = [8, 11, 14, 17, 20, 23, 26, 29]
    for i, row in enumerate(right_rows):
        team_a = teams_cycle[i * 2]
        team_b = teams_cycle[i * 2 + 1]
        ws[f"K{row}"] = team_a
        ws[f"K{row + 1}"] = team_b
        ws[f"J{row}"] = team_a

    # R16 left (col C) and right (col I)
    for row in [10, 16, 22, 28]:
        ws[f"C{row}"] = "Mexico"
        ws[f"I{row}"] = "Mexico"

    # QF left (col D) and right (col H)
    for row in [13, 25]:
        ws[f"D{row}"] = "Mexico"
        ws[f"H{row}"] = "Mexico"

    # SF
    ws["E19"] = "Mexico"
    ws["G19"] = "Mexico"

    # Champion
    ws["F19"] = champion

    # Tiebreaker
    ws["E41"] = tiebreaker

    return wb


def _wb_to_bytes(wb: openpyxl.Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Path-based fixtures (write to tmp_path so parsers can open by path)
# ---------------------------------------------------------------------------

@pytest.fixture
def valid_part1_path(tmp_path):
    wb = _make_part1_wb()
    p = tmp_path / "valid_part1.xlsx"
    wb.save(str(p))
    return p


@pytest.fixture
def valid_part2_path(tmp_path):
    wb = _make_part2_wb()
    p = tmp_path / "valid_part2.xlsx"
    wb.save(str(p))
    return p


@pytest.fixture
def part1_invalid_group_pick_path(tmp_path):
    # Group A winner set to a team not in group A
    picks = {g: (GROUPS[g][0], GROUPS[g][1]) for g in GROUPS}
    picks["A"] = ("Brazil", GROUPS["A"][1])  # Brazil is not in group A
    wb = _make_part1_wb(group_picks=picks)
    p = tmp_path / "part1_invalid_group_pick.xlsx"
    wb.save(str(p))
    return p


@pytest.fixture
def part1_duplicate_finalist_path(tmp_path):
    wb = _make_part1_wb(finalist_1="Brazil", finalist_2="Brazil")
    p = tmp_path / "part1_duplicate_finalist.xlsx"
    wb.save(str(p))
    return p


@pytest.fixture
def part1_blank_name_path(tmp_path):
    wb = _make_part1_wb(name="")
    p = tmp_path / "part1_blank_name.xlsx"
    wb.save(str(p))
    return p


@pytest.fixture
def unknown_sheet_path(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "SomethingElse"
    p = tmp_path / "unknown.xlsx"
    wb.save(str(p))
    return p
