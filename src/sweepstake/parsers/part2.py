# Parser for Part 2 bracket spreadsheets (WorldCupSweep_Part2_Bracket.xlsx).
# Reads from the "Knockout Bracket" sheet using the fixed cell map in constants.py.

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

from sweepstake.constants import (
    ALL_TEAMS,
    CHAMPION_CELL,
    QF_ALL_SLOTS,
    QF_CELL_MAP,
    R16_ALL_SLOTS,
    R16_CELL_MAP,
    R32_ALL_SLOTS,
    R32_CELL_MAP,
    SF_CELL_MAP,
    SF_SLOTS,
    SHEET_PART2,
    TIEBREAKER_CELL,
)

log = logging.getLogger(__name__)


@dataclass
class Part2ParseResult:
    participant_name: str
    r32_pairs: dict[str, tuple[str, str]]   # slot -> (team_a, team_b)
    r32_winners: dict[str, str | None]
    r16_winners: dict[str, str | None]
    qf_winners: dict[str, str | None]
    sf_winners: dict[str, str | None]
    champion: str | None
    tiebreaker_final_goals: int | None
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    @property
    def is_valid(self) -> bool:
        return len(self.errors) == 0


class Part2ParseError(ValueError):
    pass


def _cell_str(ws, col: str, row: int) -> str | None:
    val = ws[f"{col}{row}"].value
    if val is None:
        return None
    return str(val).strip() or None


def _require_team(val: str | None, label: str, errors: list[str]) -> str | None:
    if val and val not in ALL_TEAMS:
        errors.append(f"{label}: '{val}' is not a recognised team.")
    return val or None


def parse(path: str | Path) -> Part2ParseResult:
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    if SHEET_PART2 not in wb.sheetnames:
        wb.close()
        raise Part2ParseError(f"Sheet '{SHEET_PART2}' not found in {Path(path).name}")
    ws = wb[SHEET_PART2]

    errors: list[str] = []
    warnings: list[str] = []

    # Participant name (B4 is shared across both templates)
    name_val = ws["B4"].value
    participant_name = str(name_val).strip() if name_val else ""
    if not participant_name:
        errors.append("Participant name (cell B4) is missing.")

    # R32 pairs and winners
    r32_pairs: dict[str, tuple[str, str]] = {}
    r32_winners: dict[str, str | None] = {}

    all_team_cells_blank = True

    for slot in R32_ALL_SLOTS:
        cell_map = R32_CELL_MAP[slot]
        col_a, row_a = cell_map["team_cells"][0]
        col_b, row_b = cell_map["team_cells"][1]
        team_a = _cell_str(ws, col_a, row_a)
        team_b = _cell_str(ws, col_b, row_b)

        if team_a or team_b:
            all_team_cells_blank = False

        _require_team(team_a, f"R32 {slot} team A", errors)
        _require_team(team_b, f"R32 {slot} team B", errors)

        r32_pairs[slot] = (team_a or "", team_b or "")

        w_col, w_row = cell_map["winner_cell"]
        winner_raw = _cell_str(ws, w_col, w_row)
        winner = _require_team(winner_raw, f"R32 {slot} winner", errors)

        # Warn if winner is not one of the two competing teams (but still accept)
        if winner and team_a and team_b and winner not in (team_a, team_b):
            warnings.append(
                f"R32 {slot}: winner '{winner}' is not one of the two teams "
                f"('{team_a}' vs '{team_b}'). Accepted anyway."
            )

        r32_winners[slot] = winner

    if all_team_cells_blank:
        errors.append(
            "All R32 team cells are blank. The bracket must be filled in before uploading."
        )

    # R16 winners
    r16_winners: dict[str, str | None] = {}
    for slot in R16_ALL_SLOTS:
        col, row = R16_CELL_MAP[slot]
        raw = _cell_str(ws, col, row)
        r16_winners[slot] = _require_team(raw, f"R16 {slot}", errors)

    # QF winners
    qf_winners: dict[str, str | None] = {}
    for slot in QF_ALL_SLOTS:
        col, row = QF_CELL_MAP[slot]
        raw = _cell_str(ws, col, row)
        qf_winners[slot] = _require_team(raw, f"QF {slot}", errors)

    # SF winners
    sf_winners: dict[str, str | None] = {}
    for slot in SF_SLOTS:
        col, row = SF_CELL_MAP[slot]
        raw = _cell_str(ws, col, row)
        sf_winners[slot] = _require_team(raw, f"SF {slot}", errors)

    # Champion
    champ_col, champ_row = CHAMPION_CELL
    champ_raw = _cell_str(ws, champ_col, champ_row)
    champion = _require_team(champ_raw, "Champion", errors)

    # Tiebreaker (must be a non-negative integer or blank)
    tb_col, tb_row = TIEBREAKER_CELL
    tb_val = ws[f"{tb_col}{tb_row}"].value
    tiebreaker: int | None = None
    if tb_val is not None:
        try:
            tiebreaker = int(tb_val)
            if tiebreaker < 0:
                errors.append(f"Tiebreaker goals must be a non-negative integer, got {tiebreaker}.")
                tiebreaker = None
        except (TypeError, ValueError):
            errors.append(
                f"Tiebreaker goals (cell E41) must be a whole number, got '{tb_val}'."
            )

    wb.close()

    return Part2ParseResult(
        participant_name=participant_name,
        r32_pairs=r32_pairs,
        r32_winners=r32_winners,
        r16_winners=r16_winners,
        qf_winners=qf_winners,
        sf_winners=sf_winners,
        champion=champion,
        tiebreaker_final_goals=tiebreaker,
        warnings=warnings,
        errors=errors,
    )
