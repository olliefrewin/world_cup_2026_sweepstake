# Parser for Part 1 entry spreadsheets (WorldCupSweep_Part1_Entry.xlsx).
# Reads from the "My Predictions" sheet using the fixed cell map in constants.py.

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

from sweepstake.constants import (
    ALL_TEAMS,
    GROUPS,
    PART1_FINALIST_1_CELL,
    PART1_FINALIST_2_CELL,
    PART1_GOLDEN_BOOT_CELL,
    PART1_GROUP_CELL_MAP,
    PART1_NAME_CELL,
    PART1_SUBMITTED_ON_CELL,
    PART1_WINNER_CELL,
    SHEET_PART1,
)

log = logging.getLogger(__name__)


@dataclass
class Part1ParseResult:
    participant_name: str
    submitted_on: str | None
    group_picks: dict[str, tuple[str, str]]  # "A" -> (winner, runner_up)
    finalist_1: str | None
    finalist_2: str | None
    winner: str | None
    golden_boot_raw: str
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    @property
    def is_valid(self) -> bool:
        return len(self.errors) == 0


class Part1ParseError(ValueError):
    pass


def _cell_value(ws, cell_ref: str) -> str | None:
    # openpyxl cell_ref like "B4" -- returns stripped string or None
    val = ws[cell_ref].value
    if val is None:
        return None
    return str(val).strip() or None


def parse(path: str | Path) -> Part1ParseResult:
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    if SHEET_PART1 not in wb.sheetnames:
        wb.close()
        raise Part1ParseError(f"Sheet '{SHEET_PART1}' not found in {Path(path).name}")
    ws = wb[SHEET_PART1]

    errors: list[str] = []
    warnings: list[str] = []

    # Participant name
    name_raw = _cell_value(ws, PART1_NAME_CELL)
    if not name_raw:
        errors.append("Participant name (cell B4) is missing.")
    participant_name = name_raw or ""

    submitted_on = _cell_value(ws, PART1_SUBMITTED_ON_CELL)

    # Group picks
    group_picks: dict[str, tuple[str, str]] = {}
    for group, cells in PART1_GROUP_CELL_MAP.items():
        winner_raw = _cell_value(ws, cells["winner"])
        runner_raw = _cell_value(ws, cells["runner_up"])

        winner = winner_raw or ""
        runner_up = runner_raw or ""

        group_teams = GROUPS[group]

        if winner and winner not in group_teams:
            errors.append(
                f"Group {group} winner '{winner}' is not in group {group} "
                f"(valid teams: {', '.join(group_teams)})."
            )
        if runner_up and runner_up not in group_teams:
            errors.append(
                f"Group {group} runner-up '{runner_up}' is not in group {group} "
                f"(valid teams: {', '.join(group_teams)})."
            )
        if winner and runner_up and winner == runner_up:
            errors.append(
                f"Group {group}: winner and runner-up cannot be the same team ('{winner}')."
            )

        group_picks[group] = (winner, runner_up)

    # Finalists
    finalist_1 = _cell_value(ws, PART1_FINALIST_1_CELL)
    finalist_2 = _cell_value(ws, PART1_FINALIST_2_CELL)

    for label, val in [("Finalist #1", finalist_1), ("Finalist #2", finalist_2)]:
        if val and val not in ALL_TEAMS:
            errors.append(f"{label} '{val}' is not a recognised team.")

    # Winner
    winner = _cell_value(ws, PART1_WINNER_CELL)
    if winner and winner not in ALL_TEAMS:
        errors.append(f"World Cup winner '{winner}' is not a recognised team.")

    # Golden Boot -- free text, any value accepted
    golden_boot_raw = _cell_value(ws, PART1_GOLDEN_BOOT_CELL) or ""
    if not golden_boot_raw:
        warnings.append("Golden Boot pick is blank -- no points available for that category.")

    wb.close()

    return Part1ParseResult(
        participant_name=participant_name,
        submitted_on=submitted_on,
        group_picks=group_picks,
        finalist_1=finalist_1,
        finalist_2=finalist_2,
        winner=winner,
        golden_boot_raw=golden_boot_raw,
        warnings=warnings,
        errors=errors,
    )
